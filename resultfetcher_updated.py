from tkinter import *
import tkinter as tk
from tkinter import ttk, messagebox
import pymysql

from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import openpyxl
import time
import matplotlib.pyplot as plt
import numpy as np

#---------------------------------------------------------------Login Function --------------------------------------
def clear():
	userentry.delete(0,END)
	passentry.delete(0,END)

def close():
	win.destroy()	


def login():
	
	#------------------ To Show Teacher data in GUI ----------------------------------
	con = pymysql.connect(host="localhost",user="root",password="",database="svit")
	cur = con.cursor()
	cur.execute("select * from teachersdata where username =%s",(user_name.get()))
	global out
	out = cur.fetchall()
	#---------------------------------------------------------------------------------
	
	if user_name.get()=="" or password.get()=="":
		messagebox.showerror("Error","Enter User Name And Password",parent=win)	
	else:
		try:
			con = pymysql.connect(host="localhost",user="root",password="",database="svit")
			cur = con.cursor()

			cur.execute("select * from teachersdata where username=%s and password = %s",(user_name.get(),password.get()))
			row = cur.fetchone()

			if row==None:
				messagebox.showerror("Error" , "Invalid User Name And Password", parent = win)

			else:
				messagebox.showinfo("Success" , "Successfully Login" , parent = win)
				close()
				dashboard()
				con.close()
		except Exception as es:
			messagebox.showerror("Error" , f"Error Dui to : {str(es)}", parent = win)

#---------------------------------------------------------------End Login Function ---------------------------------

#------------------------------------ Fetcher Code ------------------------------------------------------------------

def dashboard():
	jntua = webdriver.Chrome(executable_path="webdrivers/chromedriver.exe")


	def fetch():
		xl = openpyxl.Workbook()

		sheet = xl.active

		# ----- year -------

		b_year = year.get()
		# -----------------

		# ------ sem -------

		b_sem = sem.get()

		# -----------------

		# ----- branch -----

		b_branch = branch.get()
		# -----------------

		# ------------ Subjects -----------------

		subject=[]

		# ---------------------------------------

		getsubflag=0

		link = url_link.get()

		rline = 3

		start_hall = hall_ticket = start_hallticket.get()
		end_hall = end_hallticket.get()
		start_limit = int(start_hall[8:])
		end_limit = int(end_hall[8:])
		limit = end_limit - start_limit + 1
		code = int(hall_ticket[7:])

		

		for i in range(limit):

			hall_ticket = hall_ticket[:7] + str(code)     #last 3 digits for Incrementing

			jntua.get(link)
			jntua.find_element_by_class_name('txt').send_keys(hall_ticket)
			jntua.find_element_by_class_name('ci').click()

			code += 1
			try:
				jntua.implicitly_wait(10)
				jntua.find_element_by_xpath('/html/body/div/div[1]/div/div/center/div[1]/table')

			except NoSuchElementException:
				continue

			sheet.cell(row=rline, column=1).value = hall_ticket

			rows = len(jntua.find_elements_by_xpath(
				'//*[@id="rs"]/table/tbody/tr'))  # len of rows in tables also known as Subject length if "rows-2"
			cols = len(jntua.find_elements_by_xpath('//*[@id="rs"]/table/tbody/tr[1]/th'))  # len of the columns

			cline = 2
			
			flist=[]



			for r in range(2, rows):
				for c in range(2, 7):
					if c==2:
						if getsubflag==0:
							subname=jntua.find_element_by_xpath("//*[@id='rs']/table/tbody/tr[" + str(r) + "]/td[" + str(c) + "]").text
							tlist=list(subname.split(" "))
							fsubname=""
							for i in tlist:
								if i.upper()=="LAB":
									fsubname=fsubname+"LAB"
								else:
									fsubname=fsubname+i[0]
							subject.append(fsubname)
						continue
					if c == 5:
						continue
					data = jntua.find_element_by_xpath(
						"//*[@id='rs']/table/tbody/tr[" + str(r) + "]/td[" + str(c) + "]").text
					sheet.cell(row=rline, column=cline).value = data  # storing Marks into Excel file
					cline = cline + 1
			rline = rline + 1

			getsubflag=1

		for i in range(1, (rows - 2) * 3, 3):
			sheet.cell(row=2, column=i + 1).value = 'MID'
			sheet.cell(row=2, column=i + 2).value = 'SEM'  # sub heading of subjects
			sheet.cell(row=2, column=i + 3).value = 'P/F'

		subject.reverse()
		for i in range(1, (rows - 2) * 3, 3):
			sheet.cell(row=1, column=i + 2).value = subject.pop()  # Subjects names Entering into Excel file


		file = "results/" + str(b_year)
		fname=file + " Year " + str(b_sem) + " Sem " + b_branch + " Results.xlsx"
		xl.save(fname)     # saving excel file
		
		
		path="results\\"+str(b_year)+" Year " + str(b_sem) + " Sem " + b_branch + " Results.xlsx"
		wb_obj = openpyxl.load_workbook(path)
		sheet_obj = wb_obj.active

		m_row=sheet_obj.max_row
		m_col=sheet_obj.max_column
		total_students=m_row-2
		xlsublist=[]

		#getting subjects

		col_index=3
		while(col_index<=m_col):
			cell_obj=sheet_obj.cell(row=1,column=col_index)
			xlsublist.append(cell_obj.value)
			col_index=col_index+3

		#assigning counters to subjects

		counterlist=[]
		for i in xlsublist:
			counterlist.append(0)

		#incrementing counters based on results

		col_index=4
		for i in range(3,m_row+1):
			ci=0
			while(col_index<=m_col):
				cell_obj=sheet_obj.cell(row=i,column=col_index)
				if cell_obj.value=="P":
					counterlist[ci]=counterlist[ci]+1
				ci=ci+1
				col_index=col_index+3
			col_index=4


		# print the total number of rows and columns
		print(sheet_obj.max_row)
		print(sheet_obj.max_column)


		percentagelist=[]

		for i in counterlist:
			percentagelist.append('{:.1f}'.format((i/total_students)*100))
		print(xlsublist)
		print(percentagelist)

		#bar graph code

		plt.bar(xlsublist, counterlist, tick_label = xlsublist,width = 0.8, color = ['green'])
		plt.xlabel('Subjects')
		plt.ylabel('Pass count')
		plt.title('Results!')
		
		# Creating plot
		fig = plt.figure(figsize =(10, 7))
		p,tx,autotexts = plt.pie(counterlist, labels = xlsublist,autopct="")

		plt.legend(p,xlsublist,loc="best")

		for i,a in enumerate(autotexts):
			a.set_text("{}%".format(percentagelist[i]))
		
		# show plot
		plt.show()
		


	def jntuaresults():
		jntua.get("https://jntuaresults.ac.in/")


	gui = Tk()

	gui.geometry('1000x500+430+250')

	gui.title("SVIT JNTUA SEMESTER RESULTS AUTOMATION TOOL")
	Label(text="Sri Venkateswara Institute Of Technology (SVIT)", fg="black", font=('Comic Sans MS', 15)).place(x=300, y=1)

	# -------- Year ------

	Label(text="Year --->  ", font=13).place(x=100, y=100)
	year = IntVar()
	Radiobutton(gui, text="1", variable=year, value=1, font=10).place(x=220, y=100)
	Radiobutton(gui, text="2", variable=year, value=2, font=10).place(x=300, y=100)
	Radiobutton(gui, text="3", variable=year, value=3, font=10).place(x=380, y=100)
	Radiobutton(gui, text="4", variable=year, value=4, font=10).place(x=460, y=100)
	# --------------------

	# -------- Sem -------

	Label(text="Sem --->  ", font=13).place(x=100, y=150)
	sem = IntVar()
	Radiobutton(gui, text="1", variable=sem, value=1, font=10).place(x=220, y=150)
	Radiobutton(gui, text="2", variable=sem, value=2, font=10).place(x=300, y=150)

	# -------------------


	# ----------- Branch -------------

	branch = StringVar()
	branch.set(" ")
	Label(text="Branch --->  ", font=13).place(x=100, y=200)
	Radiobutton(gui, text="CSE", variable=branch, value="CSE", font=10).place(x=220, y=200)
	Radiobutton(gui, text="ECE", variable=branch, value="ECE", font=10).place(x=300, y=200)
	Radiobutton(gui, text="EEE", variable=branch, value="EEE", font=10).place(x=380, y=200)
	Radiobutton(gui, text="Civil", variable=branch, value="Civil", font=10).place(x=460, y=200)
	Radiobutton(gui, text="Mech", variable=branch, value="Mech", font=10).place(x=540, y=200)

	# --------------------------------

	# -------- Hallticket no. ---------

	start_hallticket = StringVar()
	end_hallticket = StringVar()

	Label(text="Hallticket No. --->", font=13).place(x=100, y=250)

	Label(text="Start : ", font=10).place(x=300, y=250)
	Entry(gui, justify="center", textvariable=start_hallticket).place(x=380, y=255)

	Label(text="End : ", font=10).place(x=520, y=250)
	Entry(gui, justify="center", textvariable=end_hallticket).place(x=600, y=255)
	# ---------------------------------

	# ------- External Entry ---------

	ext_entry = StringVar()
	Label(text="External Entry (optional) --->", font=13).place(x=100, y=300)
	Entry(gui, justify="center", textvariable=ext_entry).place(x=380, y=305)

	Button(gui, text="ADD", command=fetch).place(x=550, y=300)
	# ---------------------------------

	# ---------- Get URl -------------

	url_link = StringVar()

	Label(text="URL --->", font=13).place(x=100, y=350)
	Entry(gui, justify="center", width=50, textvariable=url_link).place(x=200, y=355)

	Button(gui, text="Get Link", command=jntuaresults).place(x=550, y=350)

	# -------------------------------

	# -------- Fetch Button ---------

	Button(gui, text="Fetch It", command=fetch).place(x=400, y=400)
	# -------------------------------

	#--------------------Teacher Profile ---------------------------------
	print(out)

	Label(text="Teacher Profile ", font=13,fg="green").place(x=780, y=60)

	Label(text="Department : ", font=13,fg="red").place(x=740, y=100)
	Label(text="User Name", font=13,fg="red").place(x=740, y=140)
	Label(text="Email :", font=13,fg="red").place(x=740, y=180)
	Label(text=out[0][0], font=13,fg="red").place(x=870, y=100)
	Label(text=out[0][1], font=13,fg="red").place(x=870, y=140)
	Label(text=out[0][3], font=13,fg="red").place(x=870, y=180)

	#---------------------------------------------------------------------

	mainloop()
	# -------------------------------------------------------------------------------------------------------------------


#------------------------------------ End Fetcher Code --------------------------------------------------------------

#----------------------------------------------------------- Signup Window --------------------------------------------------

def signup():
	# signup database connect 
	def action():
		if user_name.get()=="" or department.get()=="" or email.get()=="" or password.get()=="" or pin.get()=="":
			messagebox.showerror("Error" , "All Fields Are Required" , parent = winsignup)
		elif pin.get() != "272931":
			messagebox.showerror("Error" , "Contact Admin For security Pin" , parent = winsignup)
		else:
			try:
				con = pymysql.connect(host="localhost",user="root",password="",database="svit")
				cur = con.cursor()
				cur.execute("select * from teachersdata where username=%s",user_name.get())
				row = cur.fetchone()
				if row!=None:
					messagebox.showerror("Error" , "Staff Data Already Exits", parent = winsignup)
				else:
					cur.execute("insert into teachersdata(department,username,password,email,securitypin) values(%s,%s,%s,%s,%s)",
						(
						
						department.get(),
						user_name.get(),
						password.get(),
						email.get(),
						pin.get()
						))
					con.commit()
					con.close()
					messagebox.showinfo("Success" , "Ragistration Successfull" , parent = winsignup)
					clear()
					switch()
				
			except Exception as es:
				messagebox.showerror("Error" , f"Error Dui to : {str(es)}", parent = winsignup)

	# close signup function			
	def switch():
		winsignup.destroy()

	# clear data function
	def clear():
		user_name.delete(0,END)
		department.delete(0,END)
		email.delete(0,END)
		password.delete(0,END)
		pin.delete(0,END)


	# start Signup Window	

	winsignup = Tk()
	winsignup.title("Svit Results Fetecher")
	winsignup.maxsize(width=500 ,  height=600)
	winsignup.minsize(width=500 ,  height=600)


	#heading label
	heading = Label(winsignup , text = "Signup" , font = 'Verdana 20 bold')
	heading.place(x=80 , y=180)

	# form data label
	user_name = Label(winsignup, text= "User Name :" , font='Verdana 10 bold')
	user_name.place(x=80,y=250)

	department = Label(winsignup, text= "Department :" , font='Verdana 10 bold')
	department.place(x=80,y=280)


	email = Label(winsignup, text= "Email Address :" , font='Verdana 10 bold')
	email.place(x=80,y=320)

	password = Label(winsignup, text= "Password :" , font='Verdana 10 bold')
	password.place(x=80,y=350)

	pin = Label(winsignup, text= "Security Pin:" , font='Verdana 10 bold')
	pin.place(x=80,y=380)

	# Entry Box ------------------------------------------------------------------

	user = user_name
	user_name = StringVar()
	department = StringVar()
	email = StringVar()
	password = StringVar()
	pin = StringVar()


	user_name = Entry(winsignup, width=40 , textvariable = user_name)
	user_name.place(x=200 , y=255)


	
	department = Entry(winsignup, width=40 , textvariable = department)
	department.place(x=200 , y=285)

	
	email = Entry(winsignup, width=40,textvariable = email)
	email.place(x=200 , y=323)

	
	password = Entry(winsignup, width=40,show="*", textvariable = password)
	password.place(x=200 , y=353)

	
	pin= Entry(winsignup, width=40 ,show="*" , textvariable = pin)
	pin.place(x=200 , y=383)


	# button login and clear

	btn_signup = Button(winsignup, text = "Signup" ,font='Verdana 10 bold', command = action)
	btn_signup.place(x=200, y=413)


	btn_login = Button(winsignup, text = "Clear" ,font='Verdana 10 bold' , command = clear)
	btn_login.place(x=280, y=413)


	sign_up_btn = Button(winsignup , text="Switch To Login" , command = switch )
	sign_up_btn.place(x=350 , y =20)


	winsignup.mainloop()
#---------------------------------------------------------------------------End Singup Window-----------------------------------	


	

#------------------------------------------------------------ Login Window -----------------------------------------

win = Tk()

# app title
win.title("Svit Results Fetecher")

# window size
win.maxsize(width=500 ,  height=500)
win.minsize(width=500 ,  height=500)


#heading label
heading = Label(win , text = "Login" , font = 'Verdana 25 bold')
heading.place(x=80 , y=150)

username = Label(win, text= "User Name :" , font='Verdana 10 bold')
username.place(x=80,y=220)

userpass = Label(win, text= "Password :" , font='Verdana 10 bold')
userpass.place(x=80,y=260)

# Entry Box
user_name = StringVar()
password = StringVar()
	
userentry = Entry(win, width=40 ,textvariable = user_name)
userentry.focus()
userentry.place(x=200 , y=223)

passentry = Entry(win, width=40, show="*" ,textvariable = password)
passentry.place(x=200 , y=260)


# button login and clear

btn_login = Button(win, text = "Login" ,font='Verdana 10 bold',command = login)
btn_login.place(x=200, y=293)


btn_clear = Button(win, text = "Clear" ,font='Verdana 10 bold', command = clear)
btn_clear.place(x=260, y=293)

# signup button

sign_up_btn = Button(win , text="Switch To Sign up" , command = signup )
sign_up_btn.place(x=350 , y =20)



win.mainloop()

#-------------------------------------------------------------------------- End Login Window ---------------------------------------------------